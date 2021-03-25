import openpyxl
import pytest

from FormTestData.SignUpData import SignUpData
from PageObject.SignUpForm import SignUpForm


class TestOne():
    book = openpyxl.load_workbook("C:\\Users\\Sumit\\Desktop\\FormData.xlsx")
    sheet = book.active

    def test_sauce_success_registration(self, setUp,get_data1):
        sign_up_form = SignUpForm(self.driver)
        sign_up_form.user_name().send_keys(get_data1["UserName"])
        sign_up_form.password().send_keys(get_data1["Password"])
        sign_up_form.login_btn().click()
        web_title=sign_up_form.web_title().text
        assert web_title=="Sauce Labs Backpack","User unsuccessfully in logging in"

    def test_sauce_invalid_username(self,setUp,get_data2):
        sign_up_form = SignUpForm(self.driver)
        sign_up_form.user_name().send_keys(get_data2["UserName"])
        sign_up_form.password().send_keys(get_data2["Password"])
        sign_up_form.login_btn().click()
        error_text=sign_up_form.error_text().text
        assert "Username and password do not match any user in this service" in error_text, "Please enter valid Username"


    def test_sauce_invalid_pwd(self,setUp,get_data3):
        sign_up_form = SignUpForm(self.driver)
        sign_up_form.user_name().send_keys(get_data3["UserName"])
        sign_up_form.password().send_keys(get_data3["Password"])
        sign_up_form.login_btn().click()
        error_text = sign_up_form.error_text().text
        assert "Username and password do not match any user in this service" in error_text, "UserName and Password don't match"

    def test_sauce_empty_username(self, setUp):
        sign_up_form = SignUpForm(self.driver)
        sign_up_form.user_name()
        sign_up_form.password().send_keys(self.sheet.cell(row=2,column=3).value)
        sign_up_form.login_btn().click()
        error_text = sign_up_form.error_text().text
        assert "Username should not be empty" in error_text, "Error message given is not as per requirement"

    def test_sauce_empty_pwd(self,setUp):
        sign_up_form = SignUpForm(self.driver)
        sign_up_form.user_name().send_keys(self.sheet.cell(row=4,column=2).value)
        sign_up_form.password()
        sign_up_form.login_btn().click()
        error_text = sign_up_form.error_text().text
        assert "Password should not be empty" in error_text, "Error message given is not as per requirement"

    def test_sauce_success_empty_pwd2(self, setUp,get_data1):
        sign_up_form = SignUpForm(self.driver)
        sign_up_form.user_name().send_keys(get_data1["UserName"])
        sign_up_form.password().send_keys(get_data1["Password"])
        sign_up_form.login_btn().click()
        try:
            webtitle=sign_up_form.web_title().text
        except:
            print("User is not able to login as no such element is present")
        # else:
        #     print("Testcase fail as user is able to login successfully")
        assert webtitle =="Sauce Lab Backpack", "User has successfully logged in"


    @pytest.fixture(params=SignUpData.get_test_data1("Testcase1"))
    def get_data1(self,request):
        return request.param

    @pytest.fixture(params=SignUpData.get_test_data2("Testcase5"))
    def get_data2(self, request):
        return request.param

    @pytest.fixture(params=SignUpData.get_test_data3("Testcase2"))
    def get_data3(self, request):
        return request.param










