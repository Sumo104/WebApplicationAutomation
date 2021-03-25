import openpyxl


class SignUpData:

    @staticmethod
    def get_test_data1(TestCaseName):

        Dict={}
        book=openpyxl.load_workbook("C:\\Users\\Sumit\\Desktop\\FormData.xlsx")
        sheet=book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == TestCaseName:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]

    @staticmethod
    def get_test_data2(TestCaseName):

        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Sumit\\Desktop\\FormData.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == TestCaseName:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]

    @staticmethod
    def get_test_data3(TestCaseName):

        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Sumit\\Desktop\\FormData.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == TestCaseName:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]